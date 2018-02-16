def lambda_handler(event, context):
    import urllib.request
    import xlrd
    from openpyxl.workbook import Workbook as openpyxlWorkbook
    from openpyxl import load_workbook
    import json
    import boto3

    URL = 'https://www.iso20022.org/sites/default/files/ISO10383_MIC/ISO10383_MIC.xls/'
    DEST_FILENAME = '/tmp/file.xlsx'
    SHEET_NAME = "MICs List by CC"
    DATA_FILE = '/tmp/data.json'
    KEY_XLSX = 'file.xlsx'
    KEY = 'data.json'
    BUCKET_NAME = 'sagar-website'

    def parse_file():
        """

            parse file.xlsx and save data
            into data.json

        """

        try:
            s3 = boto3.resource('s3')
            s3.Bucket(BUCKET_NAME).download_file(KEY_XLSX, DEST_FILENAME)
            output = []
            workbook = load_workbook(DEST_FILENAME)
            worksheet = workbook[SHEET_NAME]
            headers = [i.value for i in next(worksheet.iter_rows())]
            for row in worksheet.iter_rows():
                obj = {}
                for i in range(len(headers)):
                    obj[headers[i]] = row[i].value
                output.append(obj)
            output[:] = output[1:]
            with open(DATA_FILE, 'w') as fp:
                json.dump(output, fp, indent=4)
        # return True

        except Exception as e:
            print("Error in parsing file: ", e)
            # return False


    def upload_s3():
        """

            Uploads data.json to s3 bucket
            with public access
            url: https://s3.amazonaws.com/sagar-website/data.json
            Credetntials set using aws configure in cli

        """
        try:
            s3 = boto3.resource('s3')
            bucket = s3.Bucket(BUCKET_NAME)
            resp = bucket.upload_file(
                DATA_FILE, KEY, ExtraArgs={
                    'ACL': 'public-read'})
        except Exception as e:
            print("Error in uploading file to S3: ", e)


    print("parsing xlsx file into json file")
    parse_file()
    print("uploading json file to S3 bucket")
    upload_s3()
    print("upload finished")
    resp = {
        "statusCode": 200,
        "headers": {
            "Access-Control-Allow-Origin": "*",
        },
        "body": "{'message':'Parsing completed File at https://s3.amazonaws.com/sagar-website/data.json'}"
    }
    return resp
