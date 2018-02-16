def lambda_handler(event, context):
	import urllib.request
	import xlrd
	from openpyxl.workbook import Workbook as openpyxlWorkbook
	from openpyxl import load_workbook
	import json
	import boto3

	URL = 'https://www.iso20022.org/sites/default/files/ISO10383_MIC/ISO10383_MIC.xls/'
	DEST_FILENAME = '/tmp/file.xlsx'
	SHEET_NAME = "MICs List by CC"
	DATA_FILE = '/tmp/data.json'
	KEY_XLSX = 'file.xlsx'
	KEY = 'data.json'
	BUCKET_NAME = 'sagar-website'

	def download_and_save():
	    """
	        downloading xls file from url
	        and converting it into xlsx
	        workbook and saving in file

	    """
	    try:
	        xlsBook = xlrd.open_workbook(
	            file_contents=urllib.request.urlopen(URL).read())
	        workbook = openpyxlWorkbook()
	        for i in range(0, xlsBook.nsheets):
	            xlsSheet = xlsBook.sheet_by_index(i)
	            sheet = workbook.active if i == 0 else workbook.create_sheet()
	            sheet.title = xlsSheet.name
	            for row in range(0, xlsSheet.nrows):
	                for col in range(0, xlsSheet.ncols):
	                    sheet.cell(
	                        row=row +
	                        1,
	                        column=col +
	                        1).value = xlsSheet.cell_value(
	                        row,
	                        col)

	        workbook.save(filename=DEST_FILENAME)
	    except Exception as e:
	        print("Error in downloading file: ", e)

	def upload_s3():
	    """

	        Uploads file.xls to s3 bucket

	    """
	    try:
	        s3 = boto3.resource('s3')
	        bucket = s3.Bucket(BUCKET_NAME)
	        resp = bucket.upload_file(
	            DEST_FILENAME, KEY_XLSX, ExtraArgs={
	                'ACL': 'public-read'})
	        return True
	    except Exception as e:
	        print("Error in uploading file to S3: ", e)
	        return False
	print("downloading file")
	download_and_save()
	print("uploading xlsx file to S3 bucket")
	upload_s3()
	print("upload finished")
	resp = {
    	"statusCode": 200,
    	"headers": {
        	"Access-Control-Allow-Origin": "*",
    	},
    	"body": "{'message':'Parsing completed File at https://s3.amazonaws.com/sagar-website/file.xlsx'}"
	}	
	return resp

